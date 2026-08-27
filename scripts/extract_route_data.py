from __future__ import annotations

import json
from datetime import datetime, time
from pathlib import Path

import openpyxl


UPLOADS = Path("/workspace/scratch/53e69441edda/upload")
OUTPUT = Path(__file__).resolve().parents[1] / "app" / "data" / "route-data.json"
SOURCES = [
    ("dias-uteis", UPLOADS / "ROTAS DIAS DE SEMANA AGOSTO(3)(1).xlsx"),
    ("sabado", UPLOADS / "SABADO - AGOSTO(1)(1).xlsx"),
    ("domingo", UPLOADS / "DOMINGO - AGOSTO(1)(1).xlsx"),
]


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and 0 <= value < 1:
        total_minutes = round(value * 24 * 60)
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\n", " ").split())


def looks_like_time(value: object) -> bool:
    if isinstance(value, (datetime, time)):
        return True
    if isinstance(value, (int, float)) and 0 <= value < 1:
        return True
    text = clean(value)
    if ":" not in text:
        return False
    first, _, second = text.partition(":")
    return first.isdigit() and second[:2].isdigit()


records: list[dict[str, object]] = []
summary: list[dict[str, object]] = []

for base, source in SOURCES:
    workbook = openpyxl.load_workbook(source, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        batch = "1ª entrega"
        count = 0
        for row in sheet.iter_rows(values_only=True):
            values = list(row)
            first = clean(values[0] if values else None)
            second = clean(values[1] if len(values) > 1 else None)
            joined = " ".join(clean(value) for value in values if value is not None)
            if "ENTREGA" in joined.upper() and not looks_like_time(values[0] if values else None):
                marker = next((clean(value) for value in values if "ENTREGA" in clean(value).upper()), "")
                batch = marker.lower().replace("  ", " ") or batch
                continue
            if not looks_like_time(values[0] if values else None) or not second:
                continue
            french = clean(values[2] if len(values) > 2 else None)
            milk = clean(values[3] if len(values) > 3 else None)
            notes = " | ".join(
                text
                for text in (clean(value) for value in values[4:])
                if text and text.lower() not in {"0", "0.0"}
            )
            lowered = f"{second} {notes}".lower()
            records.append(
                {
                    "id": f"{base}-{sheet_name}-{count}".lower().replace(" ", "-"),
                    "base": base,
                    "source": source.name,
                    "driver": sheet_name,
                    "batch": batch,
                    "time": clean(values[0]),
                    "client": second,
                    "french": french,
                    "milk": milk,
                    "notes": notes,
                    "rule": "sob-demanda" if any(token in lowered for token in ("se pedir", "conforme pedido", "verificar")) else "fixo" if "fixo" in lowered else "programado",
                    "registered": False,
                    "checked": False,
                }
            )
            count += 1
        summary.append({"base": base, "driver": sheet_name, "items": count})

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(
    json.dumps({"records": records, "summary": summary}, ensure_ascii=False, indent=2),
    encoding="utf-8",
)
print(f"Generated {len(records)} route records in {OUTPUT}")
