from __future__ import annotations

import json
import os
import csv
import io
import re
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import bcrypt
import jwt
import psycopg
from fastapi import Depends, FastAPI, File, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel


DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://gestao:gestao@db:5432/gestao")
JWT_SECRET = os.getenv("JWT_SECRET", "troque-esta-chave-em-producao")
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "admin@gestao.local").lower()
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "Admin@123")
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "/data/uploads"))
MAX_UPLOAD = 20 * 1024 * 1024
RELEASE_ID = os.getenv("RELEASE_ID", "2026.08.31-R3-FULLSTACK")

app = FastAPI(title="Gestão Operacional API", version="2.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[origin.strip() for origin in os.getenv("CORS_ORIGINS", "http://localhost").split(",")],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)


class LoginBody(BaseModel):
    username: str
    password: str


class PasswordBody(BaseModel):
    current_password: str
    new_password: str


class AdminUserBody(BaseModel):
    name: str
    username: str
    email: str
    role: str
    active: bool = True
    password: str | None = None


class MapPoint(BaseModel):
    latitude: float
    longitude: float


class OptimizeRouteBody(BaseModel):
    points: list[MapPoint]


def fetch_json(url: str) -> Any:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "GestaoOperacional-Panificadora/1.0 (planejamento de entregas)",
            "Accept": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=18) as response:
            return json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        raise HTTPException(502, "Serviço externo de mapas indisponível") from exc


def connection():
    return psycopg.connect(DATABASE_URL, autocommit=True)


def initialize_database() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id BIGSERIAL PRIMARY KEY,
                name VARCHAR(160) NOT NULL,
                email VARCHAR(190) UNIQUE NOT NULL,
                password_hash VARCHAR(100) NOT NULL,
                role VARCHAR(100) NOT NULL DEFAULT 'Administrador',
                active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            CREATE TABLE IF NOT EXISTS app_state (
                id SMALLINT PRIMARY KEY DEFAULT 1 CHECK (id = 1),
                payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                version BIGINT NOT NULL DEFAULT 1,
                updated_by BIGINT REFERENCES users(id),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id BIGSERIAL PRIMARY KEY,
                user_id BIGINT REFERENCES users(id),
                action VARCHAR(180) NOT NULL,
                details JSONB NOT NULL DEFAULT '{}'::jsonb,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            CREATE TABLE IF NOT EXISTS stored_files (
                id UUID PRIMARY KEY,
                original_name VARCHAR(255) NOT NULL,
                content_type VARCHAR(120),
                size_bytes BIGINT NOT NULL,
                storage_path VARCHAR(500) NOT NULL,
                uploaded_by BIGINT REFERENCES users(id),
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            """
        )
        cursor.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS username VARCHAR(100)")
        cursor.execute("UPDATE users SET username = split_part(email, '@', 1) WHERE username IS NULL OR BTRIM(username) = ''")
        # Bancos de versões antigas podem possuir usuários cujos e-mails têm o
        # mesmo prefixo (ex.: financeiro@empresaA e financeiro@empresaB). A
        # migração anterior transformava ambos em "financeiro" e o índice
        # único derrubava a API no startup. Mantemos o primeiro e tornamos os
        # demais identificadores únicos, sem apagar ou sobrescrever usuários.
        cursor.execute(
            """
            WITH ranked AS (
                SELECT id, username,
                       ROW_NUMBER() OVER (PARTITION BY LOWER(username) ORDER BY id) AS rn
                  FROM users
            )
            UPDATE users AS u
               SET username = LEFT(u.username, 82) || '-' || u.id::text
              FROM ranked AS r
             WHERE u.id = r.id AND r.rn > 1
            """
        )
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS users_username_unique ON users (LOWER(username))")
        cursor.execute("SELECT id FROM users WHERE email = %s", (ADMIN_EMAIL,))
        if cursor.fetchone() is None:
            password_hash = bcrypt.hashpw(ADMIN_PASSWORD.encode(), bcrypt.gensalt()).decode()
            cursor.execute(
                "INSERT INTO users (name, username, email, password_hash, role) VALUES (%s, %s, %s, %s, 'Administrador')",
                ("Administrador", "admin", ADMIN_EMAIL, password_hash),
            )
        cursor.execute("INSERT INTO app_state (id) VALUES (1) ON CONFLICT (id) DO NOTHING")


@app.on_event("startup")
def startup() -> None:
    initialize_database()


def make_token(user_id: int, email: str, role: str) -> str:
    now = datetime.now(timezone.utc)
    return jwt.encode({"sub": str(user_id), "email": email, "role": role, "iat": now, "exp": now + timedelta(hours=12)}, JWT_SECRET, algorithm="HS256")


def current_user(authorization: str | None = Header(default=None)) -> dict[str, Any]:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Autenticação necessária")
    try:
        claims = jwt.decode(authorization.removeprefix("Bearer "), JWT_SECRET, algorithms=["HS256"])
    except jwt.PyJWTError as exc:
        raise HTTPException(401, "Sessão inválida ou expirada") from exc
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT id, name, email, role, active FROM users WHERE id = %s", (int(claims["sub"]),))
        row = cursor.fetchone()
    if not row or not row[4]:
        raise HTTPException(401, "Usuário inativo")
    return {"id": row[0], "name": row[1], "email": row[2], "role": row[3]}


def require_admin(user: dict[str, Any] = Depends(current_user)) -> dict[str, Any]:
    if user["role"] != "Administrador":
        raise HTTPException(403, "Ação exclusiva de administrador")
    return user


def audit(user_id: int, action: str, details: dict[str, Any] | None = None) -> None:
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("INSERT INTO audit_log (user_id, action, details) VALUES (%s, %s, %s::jsonb)", (user_id, action, json.dumps(details or {})))


def normalized_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def pick(row: dict[str, Any], *names: str) -> str:
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def closing_rule_id(value: str) -> str:
    text = normalized_header(value)
    if "segunda" in text: return "closing-monday"
    if "quarta" in text: return "closing-wednesday"
    if "sexta" in text: return "closing-friday"
    if "quinz" in text or ("15" in text and ("30" in text or "ultimo" in text)): return "closing-fortnight"
    if "25" in text: return "closing-day-25"
    if "20" in text: return "closing-day-20"
    if "30" in text or "31" in text or "mensal" in text: return "closing-month-end"
    if "diar" in text: return "closing-daily-check" if "verific" in text or "pedido" in text else "closing-daily"
    return "closing-month-end"


def due_rule_id(value: str) -> str:
    text = normalized_header(value)
    if "quarta" in text: return "due-same-wednesday"
    if "10" in text and "20" in text and "30" in text: return "due-sesi"
    if "tabela" in text or "sodexo" in text: return "due-table"
    if "28" in text: return "due-28"
    if "15" in text: return "due-15"
    return "due-30"


def color_rule_id(value: str) -> str:
    text = normalized_header(value)
    if "vermel" in text: return "color-red"
    if "amarel" in text: return "color-yellow"
    return "color-green"


def map_client_row(row: dict[str, Any]) -> dict[str, Any] | None:
    name = pick(row, "cliente", "nome", "nome_razao_social", "razao_social", "empresa_cliente")
    if not name:
        return None
    closing = pick(row, "fechamento", "data_de_emissao_nota_fechamento", "data_emissao", "faturamento", "periodicidade")
    due = pick(row, "vencimento", "data_de_vencimento", "regra_de_vencimento", "prazo")
    company_value = pick(row, "empresa", "empresa_do_grupo", "contexto")
    company = "Excelência do Pão" if "excel" in normalized_header(company_value) else "Indústria de Pães Nova Esperança"
    color = pick(row, "cor", "regra_de_cor", "status", "politica")
    return {
        "name": name,
        "document": pick(row, "cnpj_cpf", "cnpj", "cpf", "documento"),
        "email": pick(row, "email", "e_mail"),
        "whatsapp": pick(row, "whatsapp", "whatssap", "telefone", "celular"),
        "company": company,
        "closing": closing,
        "closing_rule_id": closing_rule_id(closing),
        "due_rule": due,
        "due_rule_id": due_rule_id(due),
        "color_rule_id": color_rule_id(color),
        "payment": pick(row, "forma_de_pagamento", "pagamento") or "Boleto",
        "notes": pick(row, "observacoes", "observacao", "obs"),
        "requires_order_check": "pedido" in normalized_header(closing) or "verific" in normalized_header(closing),
        "associated_units": pick(row, "empresas_associadas", "unidades", "filiais"),
    }


@app.get("/api/health")
def health() -> dict[str, str]:
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT 1")
        cursor.fetchone()
    return {"status": "ok", "release": RELEASE_ID, "api": "2.1.0"}


def public_json(url: str) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "GestaoOperacional/2.0 (+consulta-cadastral)"},
    )
    try:
        with urllib.request.urlopen(request, timeout=12) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        if exc.code == 404:
            raise HTTPException(404, "Cadastro não encontrado") from exc
        raise HTTPException(502, "O serviço público de consulta não respondeu") from exc
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        raise HTTPException(502, "O serviço público de consulta não respondeu") from exc


@app.get("/api/lookup/cep/{cep}")
def lookup_cep(cep: str):
    digits = re.sub(r"\D", "", cep)
    if len(digits) != 8:
        raise HTTPException(422, "CEP inválido")
    payload = public_json(f"https://viacep.com.br/ws/{digits}/json/")
    if payload.get("erro"):
        raise HTTPException(404, "CEP não encontrado")
    return payload


@app.get("/api/lookup/cnpj/{cnpj}")
def lookup_cnpj(cnpj: str):
    digits = re.sub(r"\D", "", cnpj)
    if len(digits) != 14:
        raise HTTPException(422, "CNPJ inválido")
    return public_json(f"https://brasilapi.com.br/api/cnpj/v1/{digits}")


@app.post("/api/auth/login")
def login(body: LoginBody):
    identifier = body.username.strip().lower()
    if not identifier or not body.password:
        raise HTTPException(422, "Informe usuário/e-mail e senha")
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, name, email, password_hash, role, active
              FROM users
             WHERE LOWER(COALESCE(username, '')) = %s
                OR LOWER(email) = %s
             ORDER BY CASE WHEN LOWER(COALESCE(username, '')) = %s THEN 0 ELSE 1 END
             LIMIT 1
            """,
            (identifier, identifier, identifier),
        )
        row = cursor.fetchone()
    if not row or not row[5] or not bcrypt.checkpw(body.password.encode(), row[3].encode()):
        raise HTTPException(401, "Credenciais inválidas")
    audit(row[0], "login", {"identifier": identifier})
    return {"token": make_token(row[0], row[2], row[4]), "user": {"id": row[0], "name": row[1], "email": row[2], "role": row[4]}}


@app.post("/api/auth/change-password")
def change_password(body: PasswordBody, user: dict[str, Any] = Depends(current_user)):
    if len(body.new_password) < 8:
        raise HTTPException(422, "A nova senha deve ter ao menos 8 caracteres")
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT password_hash FROM users WHERE id = %s", (user["id"],))
        current_hash = cursor.fetchone()[0]
        if not bcrypt.checkpw(body.current_password.encode(), current_hash.encode()):
            raise HTTPException(400, "Senha atual incorreta")
        new_hash = bcrypt.hashpw(body.new_password.encode(), bcrypt.gensalt()).decode()
        cursor.execute("UPDATE users SET password_hash = %s, updated_at = NOW() WHERE id = %s", (new_hash, user["id"]))
    audit(user["id"], "change_password")
    return {"ok": True}


@app.get("/api/admin/users")
def admin_list_users(user: dict[str, Any] = Depends(require_admin)):
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT id, name, username, email, role, active, created_at, updated_at FROM users ORDER BY name")
        rows = cursor.fetchall()
    return [{"id": str(row[0]), "name": row[1], "username": row[2], "email": row[3], "role": row[4], "active": row[5], "created_at": row[6], "updated_at": row[7]} for row in rows]


@app.post("/api/admin/users")
def admin_create_user(body: AdminUserBody, user: dict[str, Any] = Depends(require_admin)):
    if not body.password or len(body.password) < 8:
        raise HTTPException(422, "A senha inicial deve ter ao menos 8 caracteres")
    password_hash = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode()
    try:
        with connection() as conn, conn.cursor() as cursor:
            cursor.execute("INSERT INTO users (name, username, email, password_hash, role, active) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id", (body.name.strip(), body.username.lower().strip(), body.email.lower().strip(), password_hash, body.role, body.active))
            user_id = cursor.fetchone()[0]
    except psycopg.errors.UniqueViolation as exc:
        raise HTTPException(409, "Já existe um usuário com este nome de usuário ou e-mail") from exc
    audit(user["id"], "admin_user_create", {"target_id": user_id, "email": body.email.lower()})
    return {"ok": True, "id": str(user_id)}


@app.put("/api/admin/users/{user_key}")
def admin_update_user(user_key: str, body: AdminUserBody, user: dict[str, Any] = Depends(require_admin)):
    with connection() as conn, conn.cursor() as cursor:
        if user_key.isdigit():
            cursor.execute("SELECT id FROM users WHERE id = %s", (int(user_key),))
        else:
            cursor.execute("SELECT id FROM users WHERE email = %s", (body.email.lower().strip(),))
        row = cursor.fetchone()
        if row is None:
            raise HTTPException(404, "Usuário de acesso ainda não existe; cadastre-o como novo")
        target_id = row[0]
        if body.password:
            if len(body.password) < 8:
                raise HTTPException(422, "A nova senha deve ter ao menos 8 caracteres")
            password_hash = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode()
            cursor.execute("UPDATE users SET name=%s, username=%s, email=%s, role=%s, active=%s, password_hash=%s, updated_at=NOW() WHERE id=%s", (body.name.strip(), body.username.lower().strip(), body.email.lower().strip(), body.role, body.active, password_hash, target_id))
        else:
            cursor.execute("UPDATE users SET name=%s, username=%s, email=%s, role=%s, active=%s, updated_at=NOW() WHERE id=%s", (body.name.strip(), body.username.lower().strip(), body.email.lower().strip(), body.role, body.active, target_id))
    audit(user["id"], "admin_user_update", {"target_id": target_id})
    return {"ok": True, "id": str(target_id)}


@app.delete("/api/admin/users/by-email/{email}")
def admin_delete_user(email: str, user: dict[str, Any] = Depends(require_admin)):
    normalized = email.lower().strip()
    if normalized == user["email"] or normalized == ADMIN_EMAIL:
        raise HTTPException(409, "A conta administrativa em uso não pode ser apagada")
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("DELETE FROM users WHERE email = %s RETURNING id", (normalized,))
        row = cursor.fetchone()
    if row is None:
        raise HTTPException(404, "Usuário não encontrado")
    audit(user["id"], "admin_user_delete", {"target_id": row[0], "email": normalized})
    return {"ok": True}


@app.get("/api/state")
def get_state(user: dict[str, Any] = Depends(current_user)):
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT payload, version, updated_at FROM app_state WHERE id = 1")
        payload, version, updated_at = cursor.fetchone()
    return {"payload": payload, "version": version, "updated_at": updated_at, "user": user}


@app.put("/api/state")
def put_state(payload: dict[str, Any], user: dict[str, Any] = Depends(current_user)):
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT payload, version FROM app_state WHERE id = 1 FOR UPDATE")
        previous_payload, previous_version = cursor.fetchone()
        previous_payload = previous_payload or {}
        changed_sections = sorted(
            key for key in set(previous_payload) | set(payload)
            if previous_payload.get(key) != payload.get(key)
        )
        cursor.execute("UPDATE app_state SET payload = %s::jsonb, version = version + 1, updated_by = %s, updated_at = NOW() WHERE id = 1 RETURNING version, updated_at", (json.dumps(payload), user["id"]))
        version, updated_at = cursor.fetchone()
    audit(user["id"], "state_update", {
        "previous_version": previous_version,
        "version": version,
        "changed_sections": changed_sections,
    })
    return {"ok": True, "version": version, "updated_at": updated_at, "changed_sections": changed_sections}


@app.get("/api/backup")
def download_backup(user: dict[str, Any] = Depends(require_admin)):
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT payload, version, updated_at FROM app_state WHERE id = 1")
        payload, version, updated_at = cursor.fetchone()
    audit(user["id"], "backup_download", {"version": version})
    return JSONResponse({"created_at": datetime.now(timezone.utc).isoformat(), "version": version, "updated_at": updated_at.isoformat(), "payload": payload}, headers={"Content-Disposition": f'attachment; filename="backup-gestao-{datetime.now():%Y%m%d-%H%M}.json"'})


@app.post("/api/backup/restore")
def restore_backup(payload: dict[str, Any], user: dict[str, Any] = Depends(require_admin)):
    restored = payload.get("payload", payload)
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("UPDATE app_state SET payload = %s::jsonb, version = version + 1, updated_by = %s, updated_at = NOW() WHERE id = 1 RETURNING version", (json.dumps(restored), user["id"]))
        version = cursor.fetchone()[0]
    audit(user["id"], "backup_restore", {"version": version})
    return {"ok": True, "version": version}


@app.get("/api/audit")
def list_audit(limit: int = 100, user: dict[str, Any] = Depends(require_admin)):
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT a.id, COALESCE(u.name, 'Sistema'), a.action, a.details, a.created_at FROM audit_log a LEFT JOIN users u ON u.id = a.user_id ORDER BY a.id DESC LIMIT %s", (min(max(limit, 1), 500),))
        rows = cursor.fetchall()
    return [{"id": row[0], "user": row[1], "action": row[2], "details": row[3], "created_at": row[4]} for row in rows]


@app.get("/api/map/geocode")
def geocode_address(q: str):
    query = q.strip()
    if len(query) < 4 or len(query) > 240:
        raise HTTPException(422, "Informe um endereço válido")

    # CEP brasileiro: resolve primeiro pelo ViaCEP e só então geocodifica o endereço completo.
    # Isso evita depender do Nominatim reconhecer uma sequência de 8 dígitos como CEP.
    digits = re.sub(r"\D", "", query)
    if len(digits) == 8 and len(re.sub(r"[\d\s.-]", "", query)) == 0:
        cep_data = public_json(f"https://viacep.com.br/ws/{digits}/json/")
        if cep_data.get("erro"):
            raise HTTPException(404, "CEP não encontrado")
        query = ", ".join(filter(None, [
            cep_data.get("logradouro"), cep_data.get("bairro"),
            cep_data.get("localidade"), cep_data.get("uf"), digits, "Brasil"
        ]))

    params = urllib.parse.urlencode(
        {
            "format": "jsonv2",
            "limit": 6,
            "countrycodes": "br",
            "addressdetails": 1,
            "q": query,
        }
    )
    payload = fetch_json(f"https://nominatim.openstreetmap.org/search?{params}")
    results = []
    for item in payload[:6]:
        try:
            results.append(
                {
                    "display_name": str(item["display_name"]),
                    "lat": float(item["lat"]),
                    "lon": float(item["lon"]),
                }
            )
        except (KeyError, TypeError, ValueError):
            continue
    return {"results": results}


@app.post("/api/map/optimize")
def optimize_route(payload: OptimizeRouteBody):
    if len(payload.points) < 2:
        raise HTTPException(422, "A rota precisa de pelo menos duas paradas")
    if len(payload.points) > 40:
        raise HTTPException(422, "Calcule no máximo 40 paradas por vez")
    for point in payload.points:
        if not (-90 <= point.latitude <= 90 and -180 <= point.longitude <= 180):
            raise HTTPException(422, "Coordenada inválida")
    coordinates = ";".join(
        f"{point.longitude:.6f},{point.latitude:.6f}" for point in payload.points
    )
    params = urllib.parse.urlencode(
        {
            "source": "first",
            "roundtrip": "false",
            "overview": "full",
            "geometries": "geojson",
            "steps": "false",
        }
    )
    data = fetch_json(
        f"https://router.project-osrm.org/trip/v1/driving/{coordinates}?{params}"
    )
    trips = data.get("trips") or []
    waypoints = data.get("waypoints") or []
    if not trips:
        raise HTTPException(422, "Não foi possível montar uma rota viária")
    trip = trips[0]
    geometry = trip.get("geometry", {}).get("coordinates", [])
    route_coordinates = [
        [float(coordinate[1]), float(coordinate[0])]
        for coordinate in geometry
        if len(coordinate) >= 2
    ]
    order = [0] * len(waypoints)
    for source_index, waypoint in enumerate(waypoints):
        destination_index = int(waypoint.get("waypoint_index", source_index))
        if 0 <= destination_index < len(order):
            order[destination_index] = source_index
    return {
        "coordinates": route_coordinates,
        "distance": round(float(trip.get("distance", 0)) / 1000, 1),
        "duration": max(1, round(float(trip.get("duration", 0)) / 60)),
        "order": order,
    }


@app.post("/api/files")
async def upload_file(file: UploadFile = File(...), user: dict[str, Any] = Depends(current_user)):
    from uuid import uuid4
    content = await file.read(MAX_UPLOAD + 1)
    if len(content) > MAX_UPLOAD:
        raise HTTPException(413, "Arquivo maior que 20 MB")
    file_id = uuid4()
    extension = Path(file.filename or "arquivo").suffix.lower()[:10]
    destination = UPLOAD_DIR / f"{file_id}{extension}"
    destination.write_bytes(content)
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("INSERT INTO stored_files (id, original_name, content_type, size_bytes, storage_path, uploaded_by) VALUES (%s, %s, %s, %s, %s, %s)", (file_id, file.filename or "arquivo", file.content_type, len(content), str(destination), user["id"]))
    audit(user["id"], "file_upload", {"file_id": str(file_id), "name": file.filename})
    return {"id": str(file_id), "name": file.filename, "size": len(content)}


@app.post("/api/import/clients")
async def import_clients(file: UploadFile = File(...), user: dict[str, Any] = Depends(current_user)):
    content = await file.read(MAX_UPLOAD + 1)
    if len(content) > MAX_UPLOAD:
        raise HTTPException(413, "Planilha maior que 20 MB")
    suffix = Path(file.filename or "").suffix.lower()
    raw_rows: list[dict[str, Any]] = []
    if suffix == ".csv":
        decoded = content.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=";,\t,")
        except csv.Error:
            dialect = csv.excel_semicolon
        reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
        raw_rows = [{normalized_header(key): value for key, value in row.items()} for row in reader]
    elif suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            values = sheet.iter_rows(values_only=True)
            headers: list[str] | None = None
            for values_row in values:
                if headers is None:
                    candidate = [normalized_header(value) for value in values_row]
                    if any(value in {"cliente", "nome", "razao_social", "nome_razao_social"} for value in candidate):
                        headers = candidate
                    continue
                row = {headers[index]: value for index, value in enumerate(values_row) if index < len(headers) and headers[index]}
                raw_rows.append(row)
    else:
        raise HTTPException(422, "Use uma planilha XLSX ou CSV")
    rows = [mapped for row in raw_rows if (mapped := map_client_row(row))]
    audit(user["id"], "client_spreadsheet_preview", {"file": file.filename, "rows": len(rows)})
    return {"rows": rows, "count": len(rows)}


@app.get("/api/files/{file_id}")
def get_file(file_id: str, user: dict[str, Any] = Depends(current_user)):
    with connection() as conn, conn.cursor() as cursor:
        cursor.execute("SELECT original_name, content_type, storage_path FROM stored_files WHERE id = %s", (file_id,))
        row = cursor.fetchone()
    if not row or not Path(row[2]).is_file():
        raise HTTPException(404, "Arquivo não encontrado")
    return FileResponse(row[2], media_type=row[1], filename=row[0])
